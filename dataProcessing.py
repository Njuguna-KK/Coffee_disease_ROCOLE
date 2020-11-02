import openpyxl
import collections
import os
from PIL import Image
import random
import json
from shutil import copyfile

IMG_DIM = 720
xlsx_path = "./Annotations/RoCoLe-classes.xlsx"
annotation_json_path = "./Annotations/RoCoLe-json.json"
photo_path_prefix = "./Photos/"
binary_path = "./binary/"
multiclass_path = "./multiclass/"

binary_classifications = {
    "healthy": 0,
    "unhealthy": 1
}

multiclass_classifications = {
    "healthy": 0,
    "rust_level_1": 1,
    "rust_level_2": 2,
    "rust_level_3": 3,
    "rust_level_4": 4,
    "red_spider_mite": 5
}

def split_into_train_val_test(dict):
    random.seed(230)

    test = []
    val = []
    train = []

    for category in dict:
        img_names = list(dict[category])
        img_names.sort()
        random.shuffle(img_names)

        test_split = int(0.1 * len(img_names))
        val_split = int(.18 * len(img_names))

        test_img_names = img_names[:test_split]
        val_img_names = img_names[test_split: test_split + val_split]
        train_img_names = img_names[test_split + val_split:]

        test.extend(test_img_names)
        val.extend(val_img_names)
        train.extend(train_img_names)

    return {
        "test": set(test),
        "val": set(val),
        "train": set(train)
    }


def generate_binary_and_multiclass_dict_old():
    wb_obj = openpyxl.load_workbook(xlsx_path)
    sheet_obj = wb_obj.active

    binary_dict = collections.defaultdict(set)
    multiclass_dict = collections.defaultdict(set)

    num_row = sheet_obj.max_row

    for i in range(2, num_row + 1):
        image_name = sheet_obj.cell(row=i, column=1).value
        binary = sheet_obj.cell(row=i, column=2).value
        multiclass = sheet_obj.cell(row=i, column=3).value

        binary_dict[binary].add(image_name)
        multiclass_dict[multiclass].add(image_name)

    return binary_dict, multiclass_dict


def generate_train_val_test_split(binary_dict, multiclass_dict):
    binary_split = split_into_train_val_test(binary_dict)
    multiclass_split = split_into_train_val_test(multiclass_dict)
    return binary_split, multiclass_split


def get_split(img, split_dict):
    if img in split_dict["test"]:
        return "test"
    if img in split_dict["val"]:
        return "val"
    return "train"


def resize_and_save(filename, output_path, size=IMG_DIM):
    """Resize the image contained in `filename` and save it to the `output_dir`"""
    image = Image.open(filename)
    # Use bilinear interpolation instead of the default "nearest neighbor" method
    image = image.resize((size, size), Image.BILINEAR)
    image.save(output_path)


def copy_photo_files_into_directories(classification_dict, new_classification_path, classification_type, split_dict):
    binary_or_multi = "binary" if "binary" in new_classification_path else "multiclass"
    for (category, images) in classification_dict.items():
        num = str(classification_type[category])
        for img in images:
            split = get_split(img, split_dict)
            make_dir(os.path.join("just_splitted", binary_or_multi, split))
            new_img_path = os.path.join("just_splitted", binary_or_multi, split, num + "_" + img)
            resize_and_save(os.path.join(photo_path_prefix, img), new_img_path)
            # copyfile(os.path.join("just_splitted", "cropped", img), new_img_path)


def categorize_train_val_test_split(verbose = False):
    (binary_dict, multiclass_dict) = generate_binary_and_multiclass_dict_old()
    if verbose:
        print("Finished categorizing pictures into their respective classes for binary and multiclass classification")
    binary_split, multiclass_split = generate_train_val_test_split(binary_dict, multiclass_dict)
    if verbose:
        print("Finished splitting dataset")
    # copy_photo_files_into_directories(binary_dict, binary_path, binary_classifications, binary_split)
    # if verbose:
    #     print("Finished copying photos into the 'binary' folder")
    copy_photo_files_into_directories(multiclass_dict, multiclass_path, multiclass_classifications, multiclass_split)
    if verbose:
        print("Finished copying photos into the 'multiclass' folder")


def make_dir(path):
    path = os.path.abspath(os.path.join(path))

    if not os.path.exists(path):
        try:
            os.makedirs(path)
        except Exception as e:
            # Raise if directory can't be made, because image cuts won't be saved.
            print('Error creating directory')
            raise e

def generate_binary_and_multiclass_dict(img_dimension = IMG_DIM):
    binary_dict = collections.defaultdict(set)
    multiclass_dict = collections.defaultdict(set)

    make_dir(os.path.join("zoom_cropped_and_splitted", "cropped"))
    with open(annotation_json_path) as json_file:
        data = json.load(json_file)
        ct = 0
        for pic_annotation in data:
            if ct % 25 == 0: print(ct)
            leaf_obj = pic_annotation["Label"]["Leaf"][0]
            geometry = leaf_obj["geometry"]
            img_name = pic_annotation["External ID"]

            binary_classif = leaf_obj["state"]
            multi_classif = pic_annotation["Label"]["classification"]
            classif_num = multiclass_classifications[multi_classif]

            image = Image.open(os.path.join(photo_path_prefix, img_name))
            width, height = image.size
            midx = int(width/2)
            midy = int(height/2)
            img_dimension_temp = img_dimension * 2
            zoom_cropped_img = image.crop((midx - img_dimension_temp, midy - img_dimension_temp, midx + img_dimension_temp, midy + img_dimension_temp))
            zoom_cropped_img_name = str(classif_num) + "_" + img_name
            zoom_cropped_img.save(os.path.join("zoom_cropped_and_splitted", "cropped", zoom_cropped_img_name))
            binary_dict[binary_classif].add(zoom_cropped_img_name)
            multiclass_dict[multi_classif].add(zoom_cropped_img_name)

            # for i in range(len(geometry)):
            #     xy = geometry[i]
            #
            #     x = xy["x"]
            #     y = xy["y"]
            #     xmin = x - img_dimension
            #     xmax = x + img_dimension
            #     ymin = y - img_dimension
            #     ymax = y + img_dimension
            #     if xmin < 0 or ymin < 0 or xmax > width or ymax > height:
            #         continue
            #
            #     new_img = image.crop((xmin, ymin, xmax, ymax))
            #     new_img_name = str(classif_num) + "_" + "{}_".format(i) + img_name
            #     new_img.save(os.path.join("cropped", new_img_name))
            #
            #     binary_dict[binary_classif].add(new_img_name)
            #     multiclass_dict[multi_classif].add(new_img_name)
            ct += 1

    return binary_dict, multiclass_dict

def main():
    categorize_train_val_test_split(True)


if __name__ == "__main__":
    main()