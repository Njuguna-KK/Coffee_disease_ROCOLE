import openpyxl
import collections
import os
from shutil import copyfile

xlsx_path = "../Annotations/RoCoLe-classes.xlsx"
photo_path_prefix = "../Photos/"
binary_path = "../binary/"
multiclass_path = "../multiclass/"

def generate_binary_and_multiclass_dict():


    wb_obj = openpyxl.load_workbook(xlsx_path)
    sheet_obj = wb_obj.active

    binary_dict = collections.defaultdict(set)
    multiclass_dict = collections.defaultdict(set)

    num_row = sheet_obj.max_row

    for i in range(2, num_row + 1):
        image_name = sheet_obj.cell(row=i, column=1).value
        binary = sheet_obj.cell(row=i, column=2).value
        multiclass = sheet_obj.cell(row=i, column=3).value

        binary_dict[binary].add(image_name)
        multiclass_dict[multiclass].add(image_name)

    return (binary_dict, multiclass_dict)

def copy_photo_files_into_directories(classification_dict, new_classification_path):
    for (category, images) in classification_dict.items():
        for img in images:
            make_dir(os.path.join(new_classification_path, category))
            copyfile(os.path.join(photo_path_prefix, img), os.path.join(new_classification_path, category, img))


def category_pictures_into_folder():
    (binary_dict, multiclass_dict) = generate_binary_and_multiclass_dict()
    copy_photo_files_into_directories(binary_dict, binary_path)
    copy_photo_files_into_directories(multiclass_dict, multiclass_path)

def make_dir(path):
    path = os.path.abspath(os.path.join(path))

    if not os.path.exists(path):
        try:
            os.makedirs(path)
        except Exception as e:
            # Raise if directory can't be made, because image cuts won't be saved.
            print('Error creating directory')
            raise e

def main():
    category_pictures_into_folder()

if __name__ == "__main__":
    main()